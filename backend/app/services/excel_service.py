import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import io
from datetime import datetime
from typing import List, Dict, Any, Tuple
from sqlalchemy.orm import Session
from app.models.question import Question, QuestionBank, QuestionType, Difficulty

TYPE_MAP_CN_TO_EN = {
    "单选题": QuestionType.SINGLE_CHOICE.value,
    "多选题": QuestionType.MULTI_CHOICE.value,
    "判断题": QuestionType.TRUE_FALSE.value,
    "填空题": QuestionType.FILL_BLANK.value,
    "问答题": QuestionType.ESSAY.value,
    "简答题": QuestionType.ESSAY.value,
}

TYPE_MAP_EN_TO_CN = {
    QuestionType.SINGLE_CHOICE.value: "单选题",
    QuestionType.MULTI_CHOICE.value: "多选题",
    QuestionType.TRUE_FALSE.value: "判断题",
    QuestionType.FILL_BLANK.value: "填空题",
    QuestionType.ESSAY.value: "问答题",
}

DIFFICULTY_MAP_CN_TO_EN = {
    "简单": "easy",
    "中等": "medium",
    "困难": "hard"
}

class ExcelService:
    @staticmethod
    def export_exam_scores_to_excel(items: List[Any], report_title: str = "考试成绩明细") -> bytes:
        """导出当前检索条件命中的成绩台账。"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "成绩明细"

        headers = [
            "答卷ID", "考试场次", "考生姓名", "账号", "邮箱", "所属部门", "作答次数",
            "客观题得分", "主观题得分", "总分", "结果", "阅卷状态", "作答用时(秒)",
            "切屏次数", "交卷时间", "完成阅卷时间",
        ]
        last_column = get_column_letter(len(headers))
        ws.merge_cells(f"A1:{last_column}1")
        ws["A1"] = report_title
        ws["A1"].font = Font(name="微软雅黑", size=16, bold=True, color="1E3A5F")
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 30
        ws.merge_cells(f"A2:{last_column}2")
        ws["A2"] = f"导出时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}    共 {len(items)} 条有效答卷（每人每场取最后一次已交卷记录）"
        ws["A2"].font = Font(name="微软雅黑", size=10, color="64748B")

        header_row = 4
        for column, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=column, value=header)
            cell.fill = PatternFill(start_color="234E70", end_color="234E70", fill_type="solid")
            cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(style="thin", color="D8E1EA"),
            right=Side(style="thin", color="D8E1EA"),
            top=Side(style="thin", color="D8E1EA"),
            bottom=Side(style="thin", color="D8E1EA"),
        )
        for row_index, item in enumerate(items, header_row + 1):
            is_graded = item.status == "GRADED"
            values = [
                item.record_id,
                item.exam_title,
                item.student_name,
                item.username,
                item.email or "",
                item.department_name,
                item.attempt_no,
                item.objective_score,
                item.subjective_score if is_graded else "",
                item.total_score if is_graded else "",
                ("及格" if item.is_passed else "未及格") if is_graded else "待出分",
                "已完成阅卷" if is_graded else "主观题待阅卷",
                item.duration_seconds,
                item.screen_switch_count,
                item.submit_time,
                item.graded_time,
            ]
            for column, value in enumerate(values, 1):
                cell = ws.cell(row=row_index, column=column, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
                cell.font = Font(name="微软雅黑", size=10, color="334155")
                if row_index % 2 == 0:
                    cell.fill = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
            for date_column in (15, 16):
                ws.cell(row=row_index, column=date_column).number_format = "yyyy-mm-dd hh:mm:ss"
            result_cell = ws.cell(row=row_index, column=11)
            if is_graded:
                color = "047857" if item.is_passed else "BE123C"
                fill = "ECFDF5" if item.is_passed else "FFF1F2"
            else:
                color, fill = "B45309", "FFFBEB"
            result_cell.font = Font(name="微软雅黑", size=10, bold=True, color=color)
            result_cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")

        widths = [10, 28, 15, 16, 24, 18, 10, 13, 13, 11, 11, 16, 15, 11, 21, 21]
        for index, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(index)].width = width
        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A{header_row}:{last_column}{max(header_row, header_row + len(items))}"
        ws.sheet_view.showGridLines = False

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def generate_template_bytes() -> bytes:
        """生成标准题库 Excel 导入模板"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "题库导入模板"

        # 表头
        headers = [
            "题型*", "题干内容*", "选项A", "选项B", "选项C", "选项D", 
            "正确答案*", "默认分值", "难度(简单/中等/困难)", "知识点标签", "答案解析"
        ]
        ws.append(headers)

        # 示例数据
        examples = [
            ["单选题", "Python 中用于定义匿名函数的关键字是？", "def", "lambda", "async", "class", "B", "5", "简单", "Python基础", "Python 使用 lambda 声明匿名函数"],
            ["多选题", "以下哪些属于 HTTP 请求中的幂等方法？", "GET", "POST", "PUT", "DELETE", "A,C,D", "5", "中等", "计算机网络", "GET/PUT/DELETE 均属于幂等方法"],
            ["判断题", "FastAPI 基于 ASGI 标准协议构建。", "", "", "", "", "正确", "5", "简单", "Web框架", "FastAPI 底层基于 Starlette，原生遵循 ASGI"],
            ["填空题", "SQL 语言中，用于按指定列排序的关键字是 ____。", "", "", "", "", "ORDER BY", "5", "简单", "数据库", "ORDER BY 用于升序或降序排序"],
            ["问答题", "请简述什么是单点登录（SSO）以及它的核心工作原理。", "", "", "", "", "用户只需登录一次即可访问多个相互信任的应用系统；核心基于统一凭据与 Token 验证", "10", "中等", "身份认证", "详见身份安全规范第4节"]
        ]
        for row in examples:
            ws.append(row)

        # 样式美化
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        for col_idx, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws.iter_rows(min_row=1, max_row=len(examples)+1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = border

        # 调整列宽
        col_widths = [12, 35, 18, 18, 18, 18, 15, 10, 15, 15, 30]
        for idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def parse_and_import_excel(file_bytes: bytes, bank_id: int, db: Session) -> Tuple[int, List[str]]:
        """解析上传的 Excel 题库文件并导入数据库"""
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active

        success_count = 0
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue

            q_type_raw = str(row[0]).strip() if row[0] else ""
            title = str(row[1]).strip() if row[1] else ""
            opt_a = str(row[2]).strip() if row[2] else ""
            opt_b = str(row[3]).strip() if row[3] else ""
            opt_c = str(row[4]).strip() if row[4] else ""
            opt_d = str(row[5]).strip() if row[5] else ""
            answer_raw = str(row[6]).strip() if row[6] else ""
            score_raw = row[7] if len(row) > 7 and row[7] is not None else 5
            diff_raw = str(row[8]).strip() if len(row) > 8 and row[8] else "中等"
            knowledge_tag = str(row[9]).strip() if len(row) > 9 and row[9] else "通用知识"
            analysis = str(row[10]).strip() if len(row) > 10 and row[10] else ""

            if not q_type_raw or not title or not answer_raw:
                errors.append(f"第 {row_idx} 行错误：题型、题干内容或正确答案不能为空")
                continue

            q_type = TYPE_MAP_CN_TO_EN.get(q_type_raw)
            if not q_type:
                errors.append(f"第 {row_idx} 行错误：不支持的题型【{q_type_raw}】")
                continue

            try:
                score = float(score_raw)
            except ValueError:
                score = 5.0

            difficulty = DIFFICULTY_MAP_CN_TO_EN.get(diff_raw, "medium")

            # 构造 options_json
            options = []
            if q_type in [QuestionType.SINGLE_CHOICE.value, QuestionType.MULTI_CHOICE.value]:
                if opt_a: options.append({"label": opt_a, "value": "A"})
                if opt_b: options.append({"label": opt_b, "value": "B"})
                if opt_c: options.append({"label": opt_c, "value": "C"})
                if opt_d: options.append({"label": opt_d, "value": "D"})

            # 处理答案 answer_json
            answer_list = []
            if q_type == QuestionType.SINGLE_CHOICE.value:
                answer_list = [answer_raw.upper().replace(" ", "")]
            elif q_type == QuestionType.MULTI_CHOICE.value:
                # 支持 "A,B,C" 或 "ABC"
                raw = answer_raw.upper().replace("，", ",").replace(" ", "")
                if "," in raw:
                    answer_list = [x.strip() for x in raw.split(",") if x.strip()]
                else:
                    answer_list = list(raw)
            elif q_type == QuestionType.TRUE_FALSE.value:
                val = "true" if answer_raw in ["正确", "对", "T", "true", "True", "1", "√"] else "false"
                answer_list = [val]
            else:
                # 填空/问答
                answer_list = [answer_raw]

            # 创建题目记录
            question = Question(
                bank_id=bank_id,
                type=q_type,
                title=title,
                options_json=json.dumps(options, ensure_ascii=False) if options else None,
                answer_json=json.dumps(answer_list, ensure_ascii=False),
                analysis=analysis,
                score=score,
                difficulty=difficulty,
                knowledge_tag=knowledge_tag
            )
            db.add(question)
            success_count += 1

        db.commit()
        return success_count, errors

    @staticmethod
    def export_questions_to_excel(questions: List[Question]) -> bytes:
        """将题库导出为 Excel 文件"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "题库导出"

        headers = [
            "ID", "题型", "题干内容", "选项A", "选项B", "选项C", "选项D", 
            "正确答案", "默认分值", "难度", "知识点标签", "答案解析"
        ]
        ws.append(headers)

        for q in questions:
            opts = json.loads(q.options_json) if q.options_json else []
            opt_map = {item.get("value"): item.get("label") for item in opts}
            ans = json.loads(q.answer_json) if q.answer_json else []
            
            ans_str = ",".join(ans)
            if q.type == QuestionType.TRUE_FALSE.value:
                ans_str = "正确" if "true" in ans else "错误"

            diff_cn = "简单" if q.difficulty == "easy" else ("困难" if q.difficulty == "hard" else "中等")

            ws.append([
                q.id,
                TYPE_MAP_EN_TO_CN.get(q.type, q.type),
                q.title,
                opt_map.get("A", ""),
                opt_map.get("B", ""),
                opt_map.get("C", ""),
                opt_map.get("D", ""),
                ans_str,
                q.score,
                diff_cn,
                q.knowledge_tag or "",
                q.analysis or ""
            ])

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
